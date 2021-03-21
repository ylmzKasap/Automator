import copy
import importlib
import os
import pprint
from pathlib import Path
import re
import shutil
import subprocess
import sys
import time

import openpyxl
import pyautogui

import keyinfo


ACTION_DURATION = 0.65  # Action time for each command

originalScreenSize = (pyautogui.size().width, pyautogui.size().height)
allCommands = []
allEpisodeNames = []
commands = []
turn = 1
error = 0  # Switches to "1" if there is an error. Prevents os.system("cls").

episodesRegex = re.compile(r"[^\d]+")

# Create projects folder.
if not os.path.exists(f"{os.getcwd()}\\projects"):
    os.mkdir(f"{os.getcwd()}\\projects")
if not os.path.exists(f"{os.getcwd()}\\sounds"):
    os.mkdir(f"{os.getcwd()}\\sounds")
if not os.path.exists(f"{os.getcwd()}\\pictures"):
    os.mkdir(f"{os.getcwd()}\\pictures")

currentProjects = [
    folder for folder in os.listdir(f"{os.getcwd()}\\projects")
    if os.path.isdir(f"{os.getcwd()}\\projects\\{folder}")
    and folder != "__pycache__"]


def correct_project_name(project_name):
    forbiddenCharacters = ["\\", "/", ":", "*", "?", "\"", "<", ">", "|", "."]  # To create a folder
    nameError = 0
    try:
        while True:
            if project_name == "" or project_name.isspace():
                project_name = pyautogui.prompt("Please provide a name for the project.")
                continue
            for character in forbiddenCharacters:
                if character in project_name:
                    project_name = pyautogui.prompt(
                        f"Project name cannot include '{character}'. Enter a new name."
                        )
                    nameError = 1
                    break
            if nameError == 1:
                nameError = 0
                continue
            break
        return project_name
    except AttributeError:
        print("\nThe program is terminated as no valid project name has been provided.")
        input("> ")
        sys.exit()


def key_to_action(keyToPress, change, insertion, imageConditional):
    global turn
    global command
    currentPosition = pyautogui.position()
    while True:
        try:
            pixelColor = pyautogui.pixel(currentPosition.x, currentPosition.y)
            break
        except OSError:
            time.sleep(0.1)
            continue

    if imageConditional == 1:
        global imageConditionalCommands
        imageConditionalCommands.append(
            [keyinfo.keyToText[keyToPress], list((currentPosition.x, currentPosition.y)), pixelColor])
        command = "icon"
        return change, insertion

    if insertion == 1:
        commands.insert(turn-1, [])
        commands[turn-1] = [keyinfo.keyToText[keyToPress]]
        commands[turn-1].append(
            list((currentPosition.x, currentPosition.y)))
        commands[turn - 1].append(pixelColor)
        turn = len(commands) + 1
        insertion = 0
        return change, insertion

    if change == 0:
        commands.append([])
    commands[turn-1] = [keyinfo.keyToText[keyToPress]]
    commands[turn-1].append(list((currentPosition.x, currentPosition.y)))
    commands[turn-1].append(pixelColor)

    if change == 1:
        change = 0
        turn = len(commands) + 1
        return change, insertion
    turn += 1
    return change, insertion


def key_to_image_action(key, imageName, change, insertion, imageConditional, clickAmount):
    global turn

    if imageConditional == 1:
        global command
        global imageConditionalCommands
        imageConditionalCommands.append(
            [keyinfo.keyToTextImage[key], f"{os.getcwd()}\\pictures\\{imageName}", clickAmount])
        command = "icon"
        return change, insertion

    if insertion == 1:
        commands.insert(turn - 1, [])
        commands[turn - 1] = [keyinfo.keyToTextImage[key]]
        commands[turn - 1].append(f"{os.getcwd()}\\pictures\\{imageName}")
        commands[turn - 1].append(clickAmount)
        turn = len(commands) + 1
        insertion = 0
        return change, insertion

    if change == 0:
        commands.append([])
    commands[turn - 1] = [keyinfo.keyToTextImage[key]]
    commands[turn - 1].append(f"{os.getcwd()}\\pictures\\{imageName}")
    commands[turn - 1].append(clickAmount)
    if change == 1:
        change = 0
        turn = len(commands) + 1
        return change, insertion
    turn += 1
    return change, insertion


def process_variable(variableDatabase):
    variableSymbol = "v"
    allowedRange = len(list(variableDatabase.keys()))
    numberError, numberErrorMessage = 0, "\nEnter a number."
    rangeError, rangeErrorMessage = 0, f"\nThe number must be between 1 and {allowedRange}."
    if allowedRange == 0:
        print("\nThere are no variables in the variable database.")
        time.sleep(2)
        return
    while True:
        variable = 0
        try:
            os.system("cls")
            if numberError == 1:
                print(numberErrorMessage)
                numberError = 0
            if rangeError == 1:
                print(rangeErrorMessage)
                rangeError = 0
            print("\nEnter a variable number.")
            print("Current variables:\n")
            for index, v in enumerate(variableDatabase.values()):
                print(f"{index+1}. {v}")
            variable = input("> ").strip()
            variable = int(variable)
        except ValueError:
            numberError = 1
            continue
        if variable < 1 or variable > allowedRange:
            rangeError = 1
            continue
        break
    variableTranslation = variableSymbol + str(variable)
    return variableTranslation


def check_recursion(commandsList):
    for i in range(len(commandsList)):
        try:
            if 'repeat_previous' in commandsList[i][0] and 'repeat_previous' in commandsList[i+1][0]:
                return True
        except IndexError:
            return False


def print_readable_commands(commandsList):
    for index, command in enumerate(commandsList):
        print(f"{index+1}. {keyinfo.format_commands(command)}")


def format_command(command):
    return keyinfo.format_commands(command)


if len(currentProjects) == 0:
    projectName = pyautogui.prompt(
        "Enter a name to start a new project."
        )
    projectName = correct_project_name(projectName)
else:
    projectName = pyautogui.prompt(
        "Choose a project:"
        "\n\nCurrent projects:\n"
        + "\n".join(currentProjects)
        + "\n\nYou can create a new project by entering a different name."
        )
    projectName = correct_project_name(projectName)

try:
    projectPath = Path.cwd() / "projects" / projectName
except TypeError:
    print("\nThe program is terminated as no valid project name has been provided.")
    input("> ")
    sys.exit()

projectPathAlternative = f"{os.getcwd()}\\projects\\{projectName}"

# Create project files for the first time.
if not projectPath.exists():
    os.makedirs(projectPath / "data")
    os.mkdir(projectPath / "search")
    newDb = openpyxl.Workbook()
    newDb.save(projectPath / "data" / "Variable Database.xlsx")
    newDb = openpyxl.Workbook()
    newDb.save(projectPath / "data" / "Wildcard Database.xlsx")
    with open(projectPath / "__init__.py", "w", encoding="utf-8") as package:
        pass
    with open(projectPath / "data" / "__init__.py", "w", encoding="utf-8") as dataPackage:
        pass
    with open(projectPath / "data" / "searchinfo.py", "w", encoding="utf-8") as searchInfo:
        searchInfo.write("assignedSearches = []")
        searchInfo.write(f"\n\ndatabaseDecision = ''")
    with open(f"{projectPath}\\projectinfo.py", "w", encoding="utf-8") as projectInfo:
        projectInfo.write(f"projectName = '{projectName}'\n")
        projectInfo.write(f"projectPath = r'{projectPathAlternative}'\n")
        projectInfo.write(f"disableImageTips = 0\n")
        projectInfo.write(f"actionTime = {ACTION_DURATION}")
    # Copy project files to project path.
    shutil.copy(f".\\projectfiles\\start.py", projectPathAlternative)
    shutil.copy(f".\\projectfiles\\data\\varsettings.py", (projectPathAlternative + "\\data"))
    shutil.copy(f".\\projectfiles\\data\\copywildcards.py", (projectPathAlternative + "\\data"))

searchImport = importlib.import_module(f"projects.{projectName}.data.searchinfo")

# Import the variables.
varDbImport = importlib.import_module(f"projects.{projectName}.data.varsettings")
if searchImport.databaseDecision == "v":
    variableDb = varDbImport.get_vars(projectPathAlternative + "\\data", "search")[0]
else:
    variableDb = varDbImport.get_vars(projectPathAlternative + "\\data", "variable")[0]

rowsOfVariables = varDbImport.get_vars(projectPathAlternative + "\\data", "variable")[1]

# Import the wildcards.
wildcardImport = importlib.import_module(f"projects.{projectName}.data.varsettings")
if searchImport.databaseDecision == "w":
    rowsOfWildcards = wildcardImport.get_vars(projectPathAlternative + "\\data", "search")[1]
else:
    rowsOfWildcards = wildcardImport.get_vars(projectPathAlternative + "\\data", "wildcard")[1]

rowsOfOriginalWildcards = wildcardImport.get_vars(projectPathAlternative + "\\data", "wildcard")[1]

# Import wildcard & variable search function
copyWildcardsImport = importlib.import_module(f"projects.{projectName}.data.copywildcards")

# Import the function which runs the commands.
runCommandsImport = importlib.import_module(f"projects.{projectName}.start")
run_commands = runCommandsImport.run_commands

# Import project info to see preferences
projectInfo = importlib.import_module(f"projects.{projectName}.projectinfo")
disableImageTips = projectInfo.disableImageTips


def disable_image_hints():
    with open(f"{projectPath}\\projectinfo.py", "w", encoding="utf-8") as projectInfo:
        projectInfo.write(f"projectName = '{projectName}'\n")
        projectInfo.write(f"projectPath = r'{projectPathAlternative}'\n")
        projectInfo.write(f"disableImageTips = 1\n")
        projectInfo.write(f"actionTime = {ACTION_DURATION}")


saveName = "savedProject.py"

# Load existing project.
if (projectPath / saveName).exists():
    savedProject = importlib.import_module(f"projects.{projectName}.savedProject")
    allCommands = savedProject.allCommandsSave
    allEpisodeNames = savedProject.allEpisodeNamesSave
    commands = copy.deepcopy(allCommands[-1])
    episode = len(allCommands)
    turn = len(commands) + 1
else:
    episode = len(allCommands) + 1

changeInPlace = 0
insertionInPlace = 0
imageConditional = 0
firstTime = 0


while True:
    if error == 0:
        os.system("cls")
    error = 0

    if firstTime == 0:
        print(f"\nProject: {projectName}")
        firstTime = 1

    if imageConditional == 0:
        try:
            print(f"\n{allEpisodeNames[episode-1]}\n")
        except IndexError:
            print(f"\n{episode}. Unnamed Episode\n")
        if len(commands) > 0:
            print_readable_commands(commands)

        if check_recursion(commands):
            print(
                "\nWARNING! There are more than one subsequent repeat assignments."
                + "\nExecuting these commands will cause a recursion error."
            )

        command = input("> ").strip()

    if command == "zzz":
        while True:
            print("\nEnter a value for the cursor to go on x coordinate.")
            x = input("> ").strip()
            print("\nEnter a value for the cursor to go on y coordinate.")
            y = input("> ").strip()
            try:
                pyautogui.moveTo(x=int(x), y=int(y), duration=1)
            except ValueError:
                os.system("cls")
                print("\nPlease enter a number.\n")
                continue
            os.system("cls")
            print(f"\nEpisode {episode}\n")
            if len(commands) > 0:
                print_readable_commands(commands)
            command = input("> ").strip()
            if command != "zzz":
                break
    if command == "z":
        try:
            pyautogui.moveTo(commands[-1][1], duration=1)
        except:
            continue
        continue
    if command == "zz":
        try:
            pyautogui.moveTo(commands[-2][1], duration=1)
        except:
            continue
        continue

    if command == "sc":
        if searchImport.databaseDecision == "":
            os.system("cls")
            print("\nDatabase decision is not defined.")
            print("First enter 'search' to make the variable adjustments.")
            error = 1
            continue
        print(copyWildcardsImport.copy_wildcards(projectPathAlternative))
        importlib.reload(searchImport)
        if searchImport.databaseDecision == "v":
            variableDb = varDbImport.get_vars(projectPathAlternative + "\\data", "search")[0]
        else:
            variableDb = varDbImport.get_vars(projectPathAlternative + "\\data", "variable")[0]
        if searchImport.databaseDecision == "w":
            rowsOfWildcards = varDbImport.get_vars(projectPathAlternative + "\\data", "search")[1]
        else:
            rowsOfWildcards = varDbImport.get_vars(projectPathAlternative + "\\data", "wildcard")[1]
        input("> ")
        continue

    if command == "help":
        os.system("cls")
        print()
        for k, v in keyinfo.helpMenu.items():
            print(f"{k}: {v}")
            error = 1
        print("\nCommand: A single automation instruction.")
        print(
            "\nEpisode: A series of commands which constitute a meaningful part of the automation process."
        )
        continue

    if command == "epi":
        os.system("cls")
        if len(allEpisodeNames) > 0:
            print("\nAll episodes:")
            print()
            for i in allEpisodeNames:
                print(i)
            print("_" * 50)
        else:
            print("\nThere are no saved episodes.")
        print()
        error = 1
        continue

    abortReplace = 0

    if command == "rep":
        if len(commands) == 0:
            print("\nThere is no command to replace.")
            time.sleep(2)
            continue
        notFound = 0
        while True:
            os.system("cls")
            if notFound == 1:
                print("\nThere is no such command.")
                notFound = 0
            print("\nAll commands:")
            print_readable_commands(commands)
            print("\nEnter a command number to replace. 'q' cancels the process.")
            allCommandsForReplacement = []
            for i in range(1, len(commands)+1):
                allCommandsForReplacement.append(str(i))
            replacedCommand = input("> ").strip()
            if replacedCommand == "q":
                abortReplace = 1
                break
            if replacedCommand not in allCommandsForReplacement:
                notFound = 1
                continue
            replacedCommand = int(replacedCommand)
            turn = replacedCommand
            changeInPlace = 1
            break

    if abortReplace == 1:
        continue

    if changeInPlace == 1 and imageConditional == 0:
        os.system("cls")
        print(f"\nEnter the command which will replace \"{format_command(commands[replacedCommand-1])}\"")
        command = input("> ").strip()
        while command not in keyinfo.allAssignments:
            os.system("cls")
            print("\nAll commands:")
            for k, v in keyinfo.allAssignmentsExplained.items():
                print(f"{k}: {v}")
            print(f"\nThere is no command called '{command}'. Please check the keys above.")
            command = input("> ").strip()

    abortInsertion = 0

    if command == "ins":
        if len(commands) == 0:
            print("\nThere is no command to insert after.")
            time.sleep(2)
            continue
        notFound = 0
        while True:
            os.system("cls")
            if notFound == 1:
                print("\nThere is no such command.")
                notFound = 0
            print("\nAll commands:")
            print_readable_commands(commands)
            print("\nEnter a command number to insert after. 'q' cancels the process.")
            allCommandsForInsertion = []
            for i in range(0, len(commands)+1):
                allCommandsForInsertion.append(str(i))
            insertionCommand = input("> ").strip()
            if insertionCommand == "q":
                abortInsertion = 1
                break
            if insertionCommand not in allCommandsForInsertion:
                notFound = 1
                continue
            insertionCommand = int(insertionCommand)
            turn = insertionCommand + 1
            insertionInPlace = 1
            break

    if abortInsertion == 1:
        continue

    if insertionInPlace == 1 and imageConditional == 0:
        os.system("cls")
        if insertionCommand == 0:
            print(
                f"\nEnter the command which will precede "
                f"\"{format_command(commands[insertionCommand])}\".")
            command = input("> ").strip()
            while command not in keyinfo.allAssignments:
                os.system("cls")
                print("\nAll commands:")
                for k, v in keyinfo.allAssignmentsExplained.items():
                    print(f"{k}: {v}")
                print(f"\nThere is no command called '{command}'. Please check the keys above.")
                command = input("> ").strip()
        else:
            print(
                f"\nEnter the command which will be executed after "
                f"\"{format_command(commands[insertionCommand-1])}\".")
            command = input("> ").strip()
            while command not in keyinfo.allAssignments:
                os.system("cls")
                print("\nAll commands:")
                for k, v in keyinfo.allAssignmentsExplained.items():
                    print(f"{k}: {v}")
                print(f"\nThere is no command called '{command}'. Please check the keys above.")
                command = input("> ").strip()

    if command == "name":
        print("\nEnter a new episode name.")
        episodeName = input("> ").strip()
        try:
            allEpisodeNames[episode-1] = f"{str(episode)}. {episodeName}"
        except IndexError:
            allEpisodeNames.append(f"{str(episode)}. {episodeName}")
        continue

    if command == "run":
        print("\nThe episode will run in 3 seconds.")
        time.sleep(3)
        try:
            run_commands(commands, ACTION_DURATION)
        except pyautogui.FailSafeException:
            pass
        except KeyboardInterrupt:
            pass
        continue

    if command == "runep":
        if len(allCommands) <= 1:
            os.system("cls")
            print("\nThere is no episode behind to run.")
            print("If you are on the second episode, please save it to run.")
            error = 1
            continue
        os.system("cls")
        print("\nAll episodes:")
        previousEpisodes = []
        for i in allEpisodeNames:
            print(i)
        print("_" * 50, "\n")
        print(f"\nRun all commands from which episode? There are {episode-1} episodes behind.\n")
        print("The current episode must be saved if it will be executed.")
        print("'q' cancels the process.")
        for i in range(1, episode):
            previousEpisodes.append(str(i))
        runEpisode = input("> ").strip()
        if runEpisode == "q":
            continue
        while runEpisode not in previousEpisodes:
            os.system("cls")
            print("\nAll episodes:")
            print()
            for i in allEpisodeNames:
                print(i)
            print(f"\nThere is no such episode. There are {episode-1} episodes behind.")
            print("Run all commands from which episode?")
            runEpisode = input("> ").strip()
        runEpisode = int(runEpisode)
        print(f"\nAll commands from episode {runEpisode} to current episode will run in 3 seconds.")
        time.sleep(3)
        try:
            for permittedCommands in allCommands[runEpisode-1:episode]:
                run_commands(permittedCommands, ACTION_DURATION)
            try:
                if commands != allCommands[episode-1]:
                    run_commands(commands, ACTION_DURATION)
            except IndexError:
                pass
        except pyautogui.FailSafeException:
            pass
        except KeyboardInterrupt:
            pass
        continue

    if command == "save":
        saveFile = open(projectPath / "savedProject.py", "w", encoding="utf8")
        saveFile.write("\nscreenSize = " + pprint.pformat(originalScreenSize) + "\n")
        try:
            allCommands[episode-1] = copy.deepcopy(commands)
        except IndexError:
            allCommands.append(copy.deepcopy(commands))
        saveFile.write("\nallCommandsSave = " + pprint.pformat(allCommands))
        commands = []
        turn = 1
        try:
            episodeName = allEpisodeNames[episode-1]
        except IndexError:
            episodeName = f"{episode}. Unnamed Episode"
        try:
            allEpisodeNames[episode-1] = episodeName
        except IndexError:
            allEpisodeNames.append(episodeName)
        saveFile.write("\n\nallEpisodeNamesSave = " + pprint.pformat(allEpisodeNames))
        episode = len(allCommands)+1
        saveFile.close()
        continue
    elif command == "saves":
        saveFile = open(projectPath / "savedProject.py", "w", encoding="utf8")
        saveFile.write("\nscreenSize = " + pprint.pformat(originalScreenSize) + "\n")
        try:
            allCommands[episode-1] = copy.deepcopy(commands)
        except IndexError:
            allCommands.append(copy.deepcopy(commands))
        saveFile.write("\nallCommandsSave = " + pprint.pformat(allCommands))
        saveFile.write("\n\nallEpisodeNamesSave = " + pprint.pformat(allEpisodeNames))
        saveFile.close()
        continue

    if command == "go":
        if len(allEpisodeNames) > len(allCommands):
            del allEpisodeNames[-1]
        if len(allCommands) == 0:
            os.system("cls")
            print("\nThere are no episodes to go.")
            error = 1
            continue
        os.system("cls")
        print("\nAll episodes:")
        print()
        for i in allEpisodeNames:
            print(i)
        if len(allCommands) == 1:
            print(f"\nThere is only {len(allCommands)} episode to go.\n")
        else:
            print(f"\n\nChoose an episode to go. There are {len(allCommands)} episodes in total.\n")
        print("Unsaved progress within current episode will be lost. 'q' cancels the process.")
        currentEpisodes = []
        for i in range(1, len(allCommands)+1):
            currentEpisodes.append(str(i))
        goEpisode = input("> ").strip()
        if goEpisode == "q":
            continue
        while goEpisode not in currentEpisodes:
            os.system("cls")
            print("\nAll episodes:")
            print()
            for i in allEpisodeNames:
                print(i)
            print(f"\nThere is no such episode. There are {len(allCommands)} episodes in total.")
            print("Choose an episode to go.")
            goEpisode = input("> ").strip()
        episode = int(goEpisode)
        commands = copy.deepcopy(allCommands[episode-1])
        turn = len(commands)+1
        continue

    if command == "copy":
        os.system("cls")
        print("\nAll episodes:")
        print()
        for i in allEpisodeNames:
            print(i)
        print(
            f"\n\nCopy which episode's content to current one? "
            + f"There are {len(allCommands)} episodes in total.\n"
            )
        copyEpisode = input("> ").strip()
        if copyEpisode == "q":
            continue
        currentEpisodes = []
        for i in range(1, len(allCommands)+1):
            currentEpisodes.append(str(i))
        while copyEpisode not in currentEpisodes:
            os.system("cls")
            print("\nAll episodes:")
            print()
            for i in allEpisodeNames:
                print(i)
            print(f"\nThere is no such episode. There are {len(allCommands)} episodes in total.")
            print("Which episode would you like to copy to the current one?")
            copyEpisode = input("> ").strip()
        copyEpisode = int(copyEpisode)
        commands = copy.deepcopy(allCommands[copyEpisode-1])
        turn = len(commands)+1
        continue

    if command == "del":
        if len(allEpisodeNames) > len(allCommands):  # Delete extra episode name if the episode is not saved
            del allEpisodeNames[-1]
        if len(allCommands) == 0:
            os.system("cls")
            print("\nThere is no episode to delete.")
            error = 1
            continue
        os.system("cls")
        print("\nAll episodes:")
        print()
        for i in allEpisodeNames:
            print(i)
        print(f"\n\nDelete which episode? There are {len(allCommands)} episodes in total.\n")
        print("Unsaved progress within current episode will be lost. 'q' cancels the process.")
        currentEpisodes = []
        for i in range(1, len(allCommands)+1):
            currentEpisodes.append(str(i))
        delEpisode = input("> ").strip()
        if delEpisode == "q":
            continue
        while delEpisode not in currentEpisodes:
            os.system("cls")
            print("\nAll episodes:")
            print()
            for i in allEpisodeNames:
                print(i)
            print(f"\nThere is no such episode. There are {len(allCommands)} episodes in total.")
            print("Delete which episode?")
            delEpisode = input("> ").strip()
        delEpisode = int(delEpisode)
        del allCommands[delEpisode-1]
        del allEpisodeNames[delEpisode-1]
        for index, i in enumerate(allEpisodeNames):
            allEpisodeNames[index] = f"{index+1}{episodesRegex.search(i).group()}"
        try:
            commands = copy.deepcopy(allCommands[len(allCommands)-1])
            episode = len(allCommands)
        except IndexError:
            commands = []
            episode = 1
        turn = len(commands)+1
        continue

    if command == "insep":
        if len(allCommands) == 0:
            print("\nThere is no episode to insert after.")
            time.sleep(2)
            continue
        notFound = 0
        while True:
            os.system("cls")
            if notFound == 1:
                print("\nThere is no such episode.")
                notFound = 0
            print("\nAll episodes:")
            for i in allEpisodeNames:
                print(i)
            print(
                    "\nAdd a new episode after which one? "
                    + "\nUnsaved progress within current episode will be lost. 'q' cancels the process."
                  )
            allEpisodesForInsertion = []
            for i in range(0, len(allCommands)+1):
                allEpisodesForInsertion.append(str(i))
            insertionCommand = input("> ").strip()
            if insertionCommand == "q":
                abortInsertion = 1
                break
            if insertionCommand not in allEpisodesForInsertion:
                notFound = 1
                continue
            insertionCommand = int(insertionCommand)
            allCommands.insert(insertionCommand, [])
            allEpisodeNames.insert(insertionCommand, [])
            allEpisodeNames[insertionCommand] = "9999. Unnamed Episode"
            for index, i in enumerate(allEpisodeNames):
                allEpisodeNames[index] = f"{index+1}{episodesRegex.search(i).group()}"
            commands = []
            episode = insertionCommand+1
            turn = len(commands)+1
            break
        continue

    if abortInsertion == 1:
        continue

    if command == "v":
        os.system("cls")
        if searchImport.databaseDecision == "v":
            variableDb = varDbImport.get_vars(projectPathAlternative + "\\data", "search")[0]
        variableToBeWritten = process_variable(variableDb)
        if variableToBeWritten is None:
            if changeInPlace == 1:
                turn = len(commands) + 1
                changeInPlace = 0
            if insertionInPlace == 1:
                turn = len(commands) + 1
                insertionInPlaceInPlace = 0
            if imageConditional == 1:
                command = "icon"
            continue
        if imageConditional == 1:
            imageConditionalCommands.append(["write_variable", variableToBeWritten])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["write_variable"]
            commands[turn-1].append(variableToBeWritten)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = [f"write_variable"]
        commands[turn-1].append(variableToBeWritten)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue
    elif command == "vdb":
        os.system("cls")
        excelPath = "C:/Program Files (x86)/Microsoft Office/root/Office16/EXCEL.exe"
        print(
            "\nSave and close the excel file after you add the variables you want."
            "\nProject variable database will be updated."
        )
        try:
            editDatabase = subprocess.Popen(
                [excelPath, f"{projectPathAlternative}\\data\\Variable Database.xlsx"]
            )
        except FileNotFoundError:
            os.system("cls")
            print("\nCould not locate the executable file of Excel.")
            print("Please go to your project's folder and launch the variable database manually.")
            print("\nPress enter to continue.")
            input("> ").strip()
            continue
        editDatabase.wait()
        varDbImport = importlib.import_module(f"projects.{projectName}.data.varsettings")
        variableDb = varDbImport.get_vars(projectPathAlternative + "\\data", "variable")[0]
        rowsOfVariables = wildcardImport.get_vars(projectPathAlternative + "\\data", "variable")[1]
        continue
    elif command == "wdb":
        os.system("cls")
        excelPath = "C:/Program Files (x86)/Microsoft Office/root/Office16/EXCEL.exe"
        print(
            "\nSave and close the excel file after you add the wildcards you want."
            "\nProject wildcard database will be updated."
        )
        try:
            editDatabase = subprocess.Popen(
                [excelPath, f"{projectPathAlternative}\\data\\Wildcard Database.xlsx"]
            )
        except FileNotFoundError:
            os.system("cls")
            print("\nCould not locate the executable file of Excel.")
            print("Please go to your project's folder and launch the wildcard database manually.")
            print("\nPress enter to continue.")
            input("> ").strip()
            continue
        editDatabase.wait()
        wildcardImport = importlib.import_module(f"projects.{projectName}.data.varsettings")
        rowsOfOriginalWildcards = wildcardImport.get_vars(projectPathAlternative + "\\data", "wildcard")[1]
        if searchImport.databaseDecision == "w":
            rowsOfWildcards = wildcardImport.get_vars(projectPathAlternative + "\\data", "search")[1]
        else:
            rowsOfWildcards = wildcardImport.get_vars(projectPathAlternative + "\\data", "wildcard")[1]
        continue

    elif command == "click":
        if imageConditional == 1:
            imageConditionalCommands.append(["blind_click"])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn - 1, [])
            commands[turn - 1] = ["blind_click"]
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn - 1] = ["blind_click"]
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "k":
        os.system("cls")
        print("\nEnter a text input.")
        writeIt = input("> ")
        if imageConditional == 1:
            imageConditionalCommands.append(["write_text", writeIt])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["write_text"]
            commands[turn-1].append(writeIt)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["write_text"]
        commands[turn-1].append(writeIt)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "web":
        os.system("cls")
        print("\nEnter the URL of a website including 'https:\\\\'.")
        webURL = input("> ").strip()
        if imageConditional == 1:
            imageConditionalCommands.append(["go_website", webURL])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["go_website"]
            commands[turn-1].append(webURL)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["go_website"]
        commands[turn-1].append(webURL)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "hot":
        os.system("cls")
        print("\nEnter a hotkey in the following format: 'ctrl shift s'")
        hotkeyDecision = input("> ").strip()
        while hotkeyDecision not in keyinfo.availableHotkeys:
            os.system("cls")
            print(f"\nThere is no hotkey called '{hotkeyDecision}'.")
            print("\nSome example hotkeys:")
            for h in keyinfo.hotkeyExamples:
                print(h)
            hotkeyDecision = input("> ").strip()
        if imageConditional == 1:
            imageConditionalCommands.append(["hotkey", hotkeyDecision])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["hotkey"]
            commands[turn-1].append(hotkeyDecision)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["hotkey"]
        commands[turn-1].append(hotkeyDecision)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "hc":
        holdKeys = ["ctrl", "shift", "alt"]
        currentPos = pyautogui.position()
        print("\nPress which key before clicking here?")
        print("\nAvailable keys:")
        for i in holdKeys:
            print(i)
        holdKey = input("> ").strip()
        while holdKey not in holdKeys:
            os.system("cls")
            print("\nThere is no such key.")
            print("\nAvailable keys:\n")
            for i in holdKeys:
                print(i)
            holdKey = input("> ").strip()
        if imageConditional == 1:
            imageConditionalCommands.append(
                ["hold_click", holdKey, list((currentPos.x, currentPos.y))]
            )
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["hold_click"]
            commands[turn-1].append(holdKey)
            commands[turn-1].append(list((currentPos.x, currentPos.y)))
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["hold_click"]
        commands[turn-1].append(holdKey)
        commands[turn-1].append(list((currentPos.x, currentPos.y)))
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "p":
        os.system("cls")
        print("\nPress which key?")
        print("\nSome key examples:")
        for i in range(0, len(keyinfo.keyboardKeysExamples), 3):
            try:
                print(
                    f"{keyinfo.keyboardKeysExamples[i]}, "
                    + f"{keyinfo.keyboardKeysExamples[i + 1]}, "
                    + f"{keyinfo.keyboardKeysExamples[i + 2]}"
                )
            except IndexError:
                pass
        keyDecision = input("> ").strip()
        while keyDecision not in list(keyinfo.availableKeyboardKeys):
            os.system("cls")
            print(f"\nThere is no such key labeled as {keyDecision}.")
            print("\nSome key examples:")
            for i in range(0, len(keyinfo.keyboardKeysExamples), 3):
                try:
                    print(
                        f"{keyinfo.keyboardKeysExamples[i]}, "
                        + f"{keyinfo.keyboardKeysExamples[i + 1]}, "
                        + f"{keyinfo.keyboardKeysExamples[i + 2]}"
                    )
                except IndexError:
                    pass
            keyDecision = input("> ").strip()
        if imageConditional == 1:
            imageConditionalCommands.append(["press_key", keyDecision])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = (["press_key"])
            commands[turn-1].append(keyDecision)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["press_key"]
        commands[turn-1].append(keyDecision)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "maxW":
        print("Current window will be maximized in 3 seconds.")
        time.sleep(3)
        activeWindow = pyautogui.getActiveWindow()
        activeWindow.maximize()
        continue
    elif command == "max":
        if imageConditional == 1:
            imageConditionalCommands.append(["maximize_window"])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["maximize_window"]
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["maximize_window"]
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "h":
        os.system("cls")
        while True:
            print("\nHold left click for how many seconds?")
            holdTime = input("> ").strip()
            try:
                holdTime = int(holdTime)
                if holdTime < 1:
                    os.system("cls")
                    print("\nEnter a positive value.")
                    continue
                break
            except ValueError:
                os.system("cls")
                print("\nEnter a number.")
                holdTime = input("> ").strip()
        if imageConditional == 1:
            imageConditionalCommands.append(["hold_mouse", holdTime])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["hold_mouse"]
            commands[turn-1].append(holdTime)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["hold_mouse"]
        commands[turn-1].append(holdTime)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "w":
        os.system("cls")
        while True:
            print("\nWait for how many seconds?")
            waiting = input("> ").strip()
            try:
                waiting = int(waiting)
                if waiting < 1:
                    os.system("cls")
                    print("\nEnter a positive value.")
                    continue
                break
            except ValueError:
                os.system("cls")
                print("\nEnter a number.")
                waiting = input("> ").strip()
        if imageConditional == 1:
            imageConditionalCommands.append(["wait", waiting])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["wait"]
            commands[turn-1].append(waiting)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["wait"]
        commands[turn-1].append(waiting)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue
    elif command == "wRandom":
        os.system("cls")
        randomWaitingStartTime = randomWaitingEndTime = 0
        print("\nEnter two values as seconds separated by a space which will mark"
              + " the start and end points of the random waiting time.")
        while True:
            print("\nEntering '15 30' waits for random seconds between 15 and 30.")
            randomWaitTime = input("> ").strip()
            randomWaitList = randomWaitTime.split()
            if len(randomWaitList) != 2:
                os.system("cls")
                print("\nPlease enter two values as seconds, separated by a space.")
                continue
            try:
                randomWaitingStartTime = randomWaitList[0]
                randomWaitingEndTime = randomWaitList[1]
                randomWaitingStartTime = int(randomWaitingStartTime)
                randomWaitingEndTime = int(randomWaitingEndTime)
            except ValueError:
                os.system("cls")
                print(f"\nValues must be numbers.")
                continue
            if randomWaitingStartTime < 0 or randomWaitingEndTime < 0:
                os.system("cls")
                print("\nValues must be greater than 0.")
                continue
            break
        if imageConditional == 1:
            imageConditionalCommands.append(
                ["wait_random", randomWaitingStartTime, randomWaitingEndTime])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn - 1, [])
            commands[turn - 1] = ["wait_random"]
            commands[turn - 1].append(randomWaitingStartTime)
            commands[turn - 1].append(randomWaitingEndTime)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn - 1] = ["wait_random"]
        commands[turn - 1].append(randomWaitingStartTime)
        commands[turn - 1].append(randomWaitingEndTime)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "repeat":
        os.system("cls")
        if imageConditional == 0:
            if turn <= 1:
                os.system("cls")
                print("\nThere is no command behind to repeat.")
                error = 1
                if changeInPlace == 1:
                    changeInPlace = 0
                if insertionInPlace == 1:
                    insertionInPlace = 0
                turn = len(commands) + 1
                continue
        elif imageConditional == 1:
            if len(imageConditionalCommands) == 0:
                print("\nThere is no command behind to repeat.")
                time.sleep(3)
                command = "icon"
                continue
        while True:
            print("\nRepeat previous command how many times?")
            repeatCount = input("> ").strip()
            if repeatCount != 'infinite':
                try:
                    repeatCount = int(repeatCount)
                    if repeatCount < 1:
                        os.system("cls")
                        print("\nEnter a positive number.")
                        continue
                    break
                except ValueError:
                    os.system("cls")
                    print("\nEnter a number or enter 'infinite'.")
                    continue
            break
        while True:
            os.system("cls")
            print(
                "\nHow long should each iteration of action take?"
                + f"\nEnter a value in seconds. Default duration is {ACTION_DURATION}."
            )
            repeatInterval = input("> ").strip()
            try:
                repeatInterval = float(repeatInterval)
                if repeatInterval < 0:
                    os.system("cls")
                    print("\nEnter a positive number or a float value.")
                    continue
                break
            except ValueError:
                os.system("cls")
                print("\nEnter a number or a float value.")
                continue
        if imageConditional == 1:
            imageConditionalCommands.append(
                ["repeat_previous", repeatCount, repeatInterval])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["repeat_previous"]
            commands[turn-1].append(repeatCount)
            commands[turn - 1].append(repeatInterval)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["repeat_previous"]
        commands[turn-1].append(repeatCount)
        commands[turn - 1].append(repeatInterval)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "repeatpattern":
        os.system("cls")
        if imageConditional == 0:
            if turn <= 2:
                os.system("cls")
                print("\nThere is no pattern behind to repeat.")
                error = 1
                if changeInPlace == 1:
                    changeInPlace = 0
                if insertionInPlace == 1:
                    insertionInPlace = 0
                turn = len(commands) + 1
                continue
        elif imageConditional == 1:
            if len(imageConditionalCommands) <= 1:
                os.system("cls")
                print("\nThere is no pattern behind to repeat.")
                time.sleep(3)
                command = "icon"
                continue
        while True:
            os.system("cls")
            if imageConditional == 0:
                while True:
                    print("\nAll commands:\n")
                    print_readable_commands(commands)
                    print("\nRepeat all commands starting from which command?")
                    repeatPatternFrom = input("> ").strip()
                    allPossibleCommands = [i+1 for i in range(len(commands[:turn-1]))]
                    try:
                        repeatPatternFrom = int(repeatPatternFrom)
                        if repeatPatternFrom not in allPossibleCommands:
                            os.system("cls")
                            print(
                                "\nEnter a valid command number."
                                + f"\nThere are {len(allPossibleCommands)} commands behind.\n"
                            )
                            continue
                        break
                    except ValueError:
                        os.system("cls")
                        print("\nEnter a number.")
                        continue
            elif imageConditional == 1:
                while True:
                    print("\nAll commands:\n")
                    print_readable_commands(imageConditionalCommands)
                    print("\nRepeat all commands starting from which command?")
                    repeatPatternFrom = input("> ").strip()
                    allPossibleCommands = [i+1 for i in range(len(imageConditionalCommands))]
                    try:
                        repeatPatternFrom = int(repeatPatternFrom)
                        if repeatPatternFrom not in allPossibleCommands:
                            os.system("cls")
                            print(
                                "\nEnter a valid command number."
                                + f"\nThere are {len(allPossibleCommands)} commands behind.\n"
                            )
                            continue
                        break
                    except ValueError:
                        os.system("cls")
                        print("\nEnter a number.")
                        continue
            break
        while True:
            print("\nRepeat specified pattern how many times?")
            repeatCount = input("> ").strip()
            if repeatCount != "infinite":
                try:
                    repeatCount = int(repeatCount)
                    if repeatCount < 1:
                        os.system("cls")
                        print("\nEnter a positive number.")
                        continue
                    break
                except ValueError:
                    os.system("cls")
                    print("\nEnter a number or enter 'infinite'.")
                    continue

        if imageConditional == 1:
            imageConditionalCommands.append(
                ["repeat_pattern", repeatCount, repeatPatternFrom])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["repeat_pattern"]
            commands[turn-1].append(repeatCount)
            commands[turn - 1].append(repeatPatternFrom)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = [f"repeat_pattern"]
        commands[turn-1].append(repeatCount)
        commands[turn - 1].append(repeatPatternFrom)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "wild":
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn-1] = ["wildcard"]
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn-1] = ["wildcard"]
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "rfw":
        os.system("cls")
        wildcardError = 0

        if len(commands) < 2:
            print("\nThere is no pattern to repeat.")
            error = 1
            continue

        wildcardCount = 0
        for i in range(len(commands)):
            if commands[i][0] == "wildcard":
                wildcardCount += 1

        maxRowLength = len(max(rowsOfWildcards, key=len))

        allPossibleCommands = [str(i+1) for i in range(len(commands))]

        while True:
            abortRfw = 0
            while True:
                print()
                print_readable_commands(commands)
                if wildcardCount > maxRowLength != 0:
                    print("\nWARNING!")
                    print(f"There are more wildcards ({wildcardCount}) than columns ({maxRowLength}).")
                    print(
                        "Please make sure that wildcard count in a repetition cycle"
                        + "\ndoes not exceed the number of columns in the variable database.")

                print("\nRepeat all commands for the amount of variables starting from which command?")
                startingCommand = input("> ").strip()
                if startingCommand == "q":
                    abortRfw = 1
                    break
                if startingCommand not in allPossibleCommands:
                    os.system("cls")
                    print(f"\nThere is no command number called {startingCommand}.")
                    print(f"There are {len(allPossibleCommands)} commands in total.")
                    continue
                startingCommand = int(startingCommand)
                if startingCommand == len(commands):
                    os.system("cls")
                    print("\nCannot start from the last command.")
                    continue
                break

            if abortRfw == 1:
                break

            os.system("cls")
            while True:
                print()
                print_readable_commands(commands)
                print("\nAfter which command should the repetition stop?")
                endingCommand = input("> ").strip()
                if endingCommand not in allPossibleCommands:
                    os.system("cls")
                    print(f"\nThere is no command number called {endingCommand}.")
                    print(f"There are {len(allPossibleCommands)} commands in total.")
                    continue
                endingCommand = int(endingCommand)
                break

            if startingCommand - endingCommand == 0:
                os.system("cls")
                print("\nThere must be at least two commands between the starting and ending point.")
                continue
            break

        if abortRfw == 1:
            abortRfw = 0
            continue

        commands.insert(startingCommand - 1, ["repeat_commands_for_wildcards"])
        commands.insert(endingCommand + 1, ["end_repeat_commands_for_wildcards"])
        turn += 2
        continue

    elif command == "l":
        os.system("cls")
        abortLaunch = 0
        while True:
            print("\nEnter the full path of the file which will be launched.")
            launchFilePath = input("> ").strip()
            launchFilePath = launchFilePath.strip("\"\'")
            if launchFilePath == "q":
                abortLaunch = 1
                break
            if not os.path.exists(launchFilePath):
                os.system("cls")
                print(f"\n{launchFilePath} does not exist.")
                continue
            break

        if abortLaunch == 1:
            abortLaunch = 0
            continue

        if imageConditional == 1:
            imageConditionalCommands.append(
                ["launch", launchFilePath])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn - 1, [])
            commands[turn - 1] = ["launch"]
            commands[turn - 1].append(launchFilePath)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn - 1] = ["launch"]
        commands[turn - 1].append(launchFilePath)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "mr":
        os.system("cls")
        while True:
            print(
                "\nEnter a value to move the cursor relatively on x-axis."
                + "\n100: Go 100 pixels right."
                + "\n-100: Go 100 pixels left."
            )
            xDirection = input("> ").strip()
            try:
                xDirection = int(xDirection)
                break
            except ValueError:
                os.system("cls")
                print("\nPlease enter a number.")
                continue

        os.system("cls")
        while True:
            print(
                "\nEnter a value to move the. cursor relatively on y-axis."
                + "\n-100: Go 100 pixels up."
                + "\n100: Go 100 pixels down."
            )
            yDirection = input("> ").strip()
            try:
                yDirection = int(yDirection)
                break
            except ValueError:
                os.system("cls")
                print("\nPlease enter a number.")
                continue

        if imageConditional == 1:
            imageConditionalCommands.append(["move_relative", xDirection, yDirection])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn-1, [])
            commands[turn - 1] = ["move_relative"]
            commands[turn - 1].append(xDirection)
            commands[turn - 1].append(yDirection)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn - 1] = ["move_relative"]
        commands[turn - 1].append(xDirection)
        commands[turn - 1].append(yDirection)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command in list(keyinfo.keyToText.keys()):
        changeInPlace, insertionInPlace = key_to_action(
            command, changeInPlace, insertionInPlace, imageConditional)
        continue

    elif command == "sound":
        os.system("cls")
        print(
            "\nCopy the sound file you wish to use to the 'sounds' folder"
            + " which is located in the project directory."
        )
        rename = 0
        while True:
            if rename == 0:
                print("\nEnter the full name of the sound file, including its extension.")
                soundFileName = input("> ").strip()
                rename = 1
                os.system("cls")
            print(f"\nThe name of the sound file is saved as '{soundFileName}'.")
            print(
                "\nEnter 'b' if the sound should play while other commands are running."
                "\nEnter 'w' if the program should wait for the sound to end."
                "\nEnter 'ren' if you want to rename the sound file."
            )
            soundDecision = input("> ").strip()
            musicPlayStyle = ""
            if soundDecision == "ren":
                os.system("cls")
                rename = 0
                continue
            elif soundDecision == "b":
                musicPlayStyle = "background"
                break
            elif soundDecision == "w":
                musicPlayStyle = "wait"
                break
            else:
                os.system("cls")
                print("\nThere is no such command.")

        if imageConditional == 1:
            imageConditionalCommands.append(
                ["play_sound", f"{os.getcwd()}\\sounds\\{soundFileName}", musicPlayStyle])
            command = "icon"
            continue
        if insertionInPlace == 1:
            commands.insert(turn - 1, [])
            commands[turn - 1] = ["play_sound"]
            commands[turn - 1].append(f"{os.getcwd()}\\sounds\\{soundFileName}")
            commands[turn - 1].append(musicPlayStyle)
            turn = len(commands) + 1
            insertionInPlace = 0
            continue
        if changeInPlace == 0:
            commands.append([])
        commands[turn - 1] = ["play_sound"]
        commands[turn - 1].append(f"{os.getcwd()}\\sounds\\{soundFileName}")
        commands[turn - 1].append(musicPlayStyle)
        if changeInPlace == 1:
            turn = len(commands) + 1
            changeInPlace = 0
            continue
        turn += 1
        continue

    elif command == "i":
        os.system("cls")
        if disableImageTips == 0:
            print(
                "\nTake a screenshot of the image you wish to be found on the screen."
                "\nThen, copy the image file to 'pictures' folder which is found in the directory of this program."
                "\n\nPress enter after completing this process. Enter 'disable' to disable this tip."
            )
            imageTips = input("> ").strip()
            if imageTips == "disable":
                disableImageTips = 1
                disable_image_hints()
            os.system("cls")
        print("\nEnter the full name of the image file, including its extension.")
        ssName = input("> ").strip()
        os.system("cls")
        print(f"\nName of the image file saved as '{ssName}'.")
        while True:
            print("\nAvailable Commands:")
            for k, v in keyinfo.imageCommandsExplained.items():
                print(f"{k}: {v}")
            print("\nWhich command should be used on this image?")
            decision = input("> ").strip()
            if decision not in list(keyinfo.keyToTextImage.keys()):
                os.system("cls")
                print("\nThere is no such command.")
                continue
            break
        clickCount = 1
        if decision == "..":
            os.system("cls")
            while True:
                print("\nHow many times should be the action carried out?")
                clickCount = input("> ").strip()
                try:
                    clickCount = int(clickCount)
                    if clickCount < 1:
                        os.system("cls")
                        print("\nEnter a positive number.")
                        continue
                    break
                except ValueError:
                    os.system("cls")
                    print("\nEnter a number.")
                    continue
        changeInPlace, insertionInPlace = key_to_image_action(
            decision, ssName, changeInPlace, insertionInPlace, imageConditional, clickCount)
        continue

    elif command == "icon":
        os.system("cls")
        if imageConditional == 0:
            while True:
                print(
                    "\nEnter an image recognition condition:"
                    "\n\n'if': Execute the commands if the image is on the screen"
                    "\n'if not': Execute the commands if the image is not on the screen"
                    "\n'while': Execute the commands while the image is on the screen"
                    "\n'while not': Execute the commands while the image is not on the screen"
                )
                conditionalDecision = input("> ").strip()
                conditionalTypes = ["if", "if not", "while", "while not"]
                if conditionalDecision not in conditionalTypes:
                    os.system("cls")
                    print("\nThere is no such conditional.")
                    continue
                break
            os.system("cls")
            if disableImageTips == 0:
                print(
                    "\nTake a screenshot of the image you wish to be found on the screen."
                    "\nThen, copy the image file to 'pictures' folder which is found in the directory of this program."
                    "\n\nPress enter after completing this process. Enter 'disable' to disable this tip."
                )
                imageTips = input("> ").strip()
                if imageTips == "disable":
                    disableImageTips = 1
                    disable_image_hints()
                os.system("cls")
            print("\nEnter the full name of the image file, including its extension.")
            imageName = input("> ").strip()
            os.system("cls")
            print(f"\nName of the image file saved as '{imageName}'.")
            imageConditional = 1
            imageConditionalCommands = []
        while True:
            if len(imageConditionalCommands) > 0:
                print()
                print_readable_commands(imageConditionalCommands)
            print(
                f"\nEnter a command which will be executed {conditionalDecision} the image is on the screen."
                "\nPress 'q' to end the command assignment, '-' to delete the last command."
            )
            if check_recursion(imageConditionalCommands):
                print(
                    "\nWARNING! There are more than one subsequent repeat assignments."
                    + "\nExecuting these commands will cause a recursion error."
                )
            command = input("> ").strip()
            if command == "q":
                if len(imageConditionalCommands) > 0:
                    imageConditional = 0
                    break
                elif len(imageConditionalCommands) == 0:
                    os.system("cls")
                    print("\nEnter at least one command to quit.")
                    continue
            if command == "-":
                try:
                    os.system("cls")
                    imageConditionalCommands.pop()
                    continue
                except IndexError:
                    os.system("cls")
                    print("\nThe command list is empty.")
                    continue
            if command not in list(keyinfo.imageConditionalAssignments):
                os.system("cls")
                print("\nThere is no such command. Available commands:\n")
                for k, v in keyinfo.imageConditionalAssignmentsExplained.items():
                    print(f"{k}: {v}")
                continue
            break
        if imageConditional == 0 and len(imageConditionalCommands) > 0:
            if insertionInPlace == 1:
                commands.insert(turn - 1, [])
                commands[turn - 1] = ["image_conditional"]
                commands[turn - 1].append(f"{os.getcwd()}\\pictures\\{imageName}")
                commands[turn - 1].append(imageConditionalCommands)
                commands[turn - 1].append(conditionalDecision)
                turn = len(commands) + 1
                insertionInPlace = 0
                imageConditionalCommands = []
                continue
            if changeInPlace == 0:
                commands.append([])
            commands[turn - 1] = ["image_conditional"]
            commands[turn - 1].append(f"{os.getcwd()}\\pictures\\{imageName}")
            commands[turn - 1].append(imageConditionalCommands)
            commands[turn - 1].append(conditionalDecision)
            if changeInPlace == 1:
                turn = len(commands) + 1
                changeInPlace = 0
                imageConditionalCommands = []
                continue
            turn += 1
            imageConditionalCommands = []
            continue
        elif imageConditional == 1:
            continue

    elif command == "search":
        os.system("cls")
        abortSearch = 0
        importlib.reload(searchImport)
        assignedSearches = searchImport.assignedSearches

        if searchImport.databaseDecision == "":
            while True:
                print("\nPull variables from which database?")
                print("'w': Wildcard Database")
                print("'v': Variable Database")
                databaseDecision = input("> ").strip().lower()
                if databaseDecision not in ["w", "v"]:
                    os.system("cls")
                    print("\nEnter 'w' or 'v'.")
                    continue
                break
        else:
            databaseDecision = searchImport.databaseDecision

        # Return if the selected database is empty.
        if databaseDecision == "w":
            if len(rowsOfOriginalWildcards) == 0:
                os.system("cls")
                print("\nThere are no wildcards in the wildcard database.")
                error = 1
                continue
        elif databaseDecision == "v":
            if len(rowsOfVariables) == 0:
                os.system("cls")
                print("\nThere are no variables in the variable database.")
                error = 1
                continue

        if len(assignedSearches) >= 1:
            while True:
                if databaseDecision == "w":
                    databaseName = "wildcard database"
                elif databaseDecision == "v":
                    databaseName = "variable database"

                print("\nCurrent search assignments:")
                for i, sList in enumerate(assignedSearches):
                    print(
                        f"{i + 1}. Column {sList[2]} in {databaseName}: "
                        + f"Search '{', '.join(sList[1])}' files"
                        + f" in '{sList[0]}'"
                    )

                if databaseDecision == "w":
                    if len(rowsOfVariables) >= 1:
                        print(f"\nData will be pulled from the {databaseName}. Enter 'c' to change it.")
                if databaseDecision == "v":
                    if len(rowsOfOriginalWildcards) >= 1:
                        print(f"\nData will be pulled from the {databaseName}. Enter 'c' to change it.")

                print("\nPress '-' to delete the last assignment, enter to continue.")
                searchDecision = input("> ").strip()

                if searchDecision.lower() == "c":
                    os.system("cls")
                    if databaseDecision == "w":
                        if len(rowsOfVariables) == 0:
                            print("\nThere are no variables in the variable database.")
                            continue
                        databaseDecision = "v"
                    elif databaseDecision == "v":
                        if len(rowsOfOriginalWildcards) == 0:
                            print("\nThere are no wildcards in the wildcard database.")
                            continue
                        databaseDecision = "w"

                    with open(projectPath / "data" / "searchinfo.py", "w", encoding="utf-8") as searchInfo:
                        searchInfo.write(f"assignedSearches = {pprint.pformat(assignedSearches)}")
                        searchInfo.write(f"\n\ndatabaseDecision = '{databaseDecision}'")
                    continue

                elif searchDecision == "-":
                    try:
                        os.system("cls")
                        assignedSearches.pop()
                        continue
                    except IndexError:
                        os.system("cls")
                        continue

                elif searchDecision == "q":
                    abortSearch = 1
                break

        if abortSearch == 1:
            abortSearch = 0
            continue

        # Choose the directory to search the variables or wildcards.
        os.system("cls")
        while True:
            print("\nEnter the full path of the directory you wish to be searched.")
            print("Enter 'q' to exit search process.")
            searchDirectory = input("> ").strip(" \"\'")
            if searchDirectory == "q":
                abortSearch = 1
                break
            if not os.path.exists(searchDirectory):
                os.system("cls")
                print("\nThere is no such directory.")
                continue
            break

        if abortSearch == 1:
            abortSearch = 0
            continue

        # Choose file extensions to apply the variables or wildcards.
        os.system("cls")
        fileExtensions = []
        while True:
            print("\nEnter the file extensions which will be applied to the wildcards while searching.")
            if len(fileExtensions) > 0:
                print(f"\nDirectory '{os.path.basename(searchDirectory)}' will be searched for:")
                for i, extension in enumerate(fileExtensions):
                    print(f"{i+1}. {extension} files")
            searchedFileExtension = input("> ")
            if searchedFileExtension == "":
                os.system("cls")
                print("\nEnter 'q' to end current process.")
                continue
            searchedFileExtension = searchedFileExtension.strip()
            if searchedFileExtension == "-":
                os.system("cls")
                if len(fileExtensions) >= 1:
                    del fileExtensions[-1]
                    continue
            elif searchedFileExtension == "q":
                if len(fileExtensions) == 0:
                    os.system("cls")
                    print("\nAt least one extension is required.")
                    continue
                break
            else:
                fileExtensions.append(searchedFileExtension)

        # Calculate the maximum number of columns.
        maxRowLength = 0
        if databaseDecision == "w":
            wb = openpyxl.load_workbook(projectPath / "data" / "Wildcard Database.xlsx")
            sheet = wb.active
            maxRowLength = len(list(sheet.columns))
            dataType = "wildcard"
            wb.close()
        elif databaseDecision == "v":
            maxRowLength = len(max(rowsOfVariables, key=len))
            dataType = "variable"

        # Choose the column to apply extensions.
        os.system("cls")
        variableToApplyExtensions = 0
        if maxRowLength == 1:
            variableToApplyExtensions = 1
        else:
            while True:
                print(f"\nThere are {maxRowLength} {dataType}s in a single row.")
                print("Apply extensions to the variables in which column?")
                availableVariableColumns = [str(i) for i in range(1, maxRowLength + 1)]
                variableToApplyExtensions = input("> ").strip()
                if variableToApplyExtensions not in availableVariableColumns:
                    os.system("cls")
                    print(f"\n{variableToApplyExtensions} does not exist.")
                    continue
                variableToApplyExtensions = int(variableToApplyExtensions)
                break

        # Save search info.
        assignedSearches.append([searchDirectory, fileExtensions, variableToApplyExtensions])
        with open(projectPath / "data" / "searchinfo.py", "w", encoding="utf-8") as searchInfo:
            searchInfo.write(f"assignedSearches = {pprint.pformat(assignedSearches)}")
            searchInfo.write(f"\n\ndatabaseDecision = '{databaseDecision}'")

        # Search & copy the files.
        print(copyWildcardsImport.copy_wildcards(projectPathAlternative))
        input("> ")

        # Adjust variables & wildcards according to the database decision.
        importlib.reload(searchImport)
        if searchImport.databaseDecision == "v":
            variableDb = varDbImport.get_vars(projectPathAlternative + "\\data", "search")[0]
        else:
            variableDb = varDbImport.get_vars(projectPathAlternative + "\\data", "variable")[0]
        if searchImport.databaseDecision == "w":
            rowsOfWildcards = varDbImport.get_vars(projectPathAlternative + "\\data", "search")[1]
        else:
            rowsOfWildcards = varDbImport.get_vars(projectPathAlternative + "\\data", "wildcard")[1]
        continue

    elif command == "qq":
        break
    elif command == "--":
        if len(commands) == 0:
            os.system("cls")
            print("\nThere is no command to delete.")
            error = 1
            continue
        elif len(commands) == 1:
            turn -= 1
            del commands[-1]
            continue
        else:
            notFound = 0
            abortRemoval = 0
            while True:
                os.system("cls")
                if notFound == 1:
                    print("\nThere is no such command.")
                    notFound = 0
                print("\nAll commands:")
                print_readable_commands(commands)
                print("\nDelete which command? 'q' cancels the process.")
                allCommandsForRemoval = []
                for i in range(1, len(commands)+1):
                    allCommandsForRemoval.append(str(i))
                deletedCommand = input("> ").strip()
                if deletedCommand == "q":
                    abortRemoval = 1
                    break
                if deletedCommand not in allCommandsForRemoval:
                    notFound = 1
                    continue
                deletedCommand = int(deletedCommand)
                break
            if abortRemoval == 1:
                abortRemoval = 0
                continue

            delCo = deletedCommand - 1
            if commands[delCo][0] == "repeat_commands_for_wildcards":
                del commands[delCo]
                turn -= 1
                for i in range(1, len(commands[delCo:])):
                    if commands[delCo + i][0] == "end_repeat_commands_for_wildcards":
                        del commands[delCo + i]
                        turn -= 1
                        break
            elif commands[delCo][0] == "end_repeat_commands_for_wildcards":
                del commands[delCo]
                turn -= 1
                for i in range(1, len(commands[:delCo]) + 1):
                    if commands[delCo - i][0] == "repeat_commands_for_wildcards":
                        del commands[delCo - i]
                        turn -= 1
                        break
            else:
                del commands[delCo]
                turn -= 1
                continue

    elif command == "-":
        if len(commands) >= 1:
            if commands[-1][0] == "end_repeat_commands_for_wildcards":
                del commands[-1]
                turn -= 1
                for i in range(len(commands)-2, -1, -1):
                    if commands[i][0] == "repeat_commands_for_wildcards":
                        del commands[i]
                        turn -= 1
                        break
            else:
                turn -= 1
                del commands[-1]
                continue
    else:
        os.system("cls")
        print("\nAll commands:\n")
        for k, v in keyinfo.allAssignmentsExplained.items():
            print(f"{k}: {v}")
        print(f"\nThere is no command called '{command}'. Please check the keys above.")
        error = 1

input("\nEND")
